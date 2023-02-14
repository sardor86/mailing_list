from PyQt6.QtWidgets import QApplication, QMainWindow, QFileDialog
import sys
from openpyxl import load_workbook

from mainwindow import Ui_MainWindow
from send_message import start_send_message


class MainWindow(Ui_MainWindow):
    def __init__(self):
        self.app = QApplication(sys.argv)
        self.window = QMainWindow()

        self.excel_file = None
        self.image_file = None

    def draw_window(self):
        self.window.show()
        sys.exit(self.app.exec())

    def send_all_messages(self):
        if (self.excel_file is None or self.image_file is None) or (self.excel_file == '' or self.image_file == ''):
            if self.excel_file is None or self.excel_file == '':
                self.choice_file.setText('Вы не выбрали excel файл')

            if self.image_file is None or self.image_file == '':
                self.choice_image.setText('Вы не выбрали картину')

            return 0

        phone_numbers = []

        wb = load_workbook(self.excel_file)
        ws = wb[wb.sheetnames[0]]
        for row in ws.values:
            try:
                int(row[self.spinBox.value()])
                phone_numbers.append(row[0])
            except ValueError:
                pass

        quantity_phone_number = 0
        for phone_number in phone_numbers:
            start_send_message(phone_number, self.image_file, self.text.toPlainText())
            quantity_phone_number += 1
            self.progressBar.setValue(int(quantity_phone_number / (len(phone_numbers)/100)))
            self.progress_sent_message.setText(f'Отправлено {quantity_phone_number}')

    def choice_excel_file(self):
        file, _ = QFileDialog.getOpenFileName(None,
                                              "Выбор excel таблицы",
                                              "./",
                                              "excel (*.xlsx *.xlsm *.xlsb *.xltx)")

        self.excel_file = str(file)
        self.choice_file.setText(str(file).split("/")[-1])

    def choice_image_file(self):
        file, _ = QFileDialog.getOpenFileName(None,
                                              "Выбор картинки",
                                              "./",
                                              "Картинка (*.jpg *.png)")

        self.image_file = str(file)
        self.choice_image.setText(str(file).split("/")[-1])

    def setting(self):
        self.setupUi(self.window)
        self.send_message.clicked.connect(self.send_all_messages)
        self.choice_file.clicked.connect(self.choice_excel_file)
        self.choice_image.clicked.connect(self.choice_image_file)


if __name__ == "__main__":
    window = MainWindow()
    window.setting()

    window.draw_window()
